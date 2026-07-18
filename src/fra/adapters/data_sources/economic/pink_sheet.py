"""World Bank Pink Sheet monthly commodity workbook adapter."""

from collections.abc import Callable
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from fra.adapters.data_sources.common.http import HttpClient
from fra.adapters.data_sources.common.manifests import validate_source_manifest
from fra.domain.economic import (
    EconomicObservation,
    EconomicSeries,
    EconomicSeriesCapabilities,
    EconomicSeriesRequest,
)
from fra.domain.errors import ExternalDataInvalidError, PointInTimeUnavailableError
from fra.domain.shared import HealthState, HealthStatus
from fra.domain.sources import DataEnvelope, SourceDescriptor

DEFAULT_URL = (
    "https://thedocs.worldbank.org/en/doc/5d903e848db1d1b83e0ec8f744e55570-"
    "0350012021/related/CMO-Historical-Data-Monthly.xlsx"
)


class WorldBankPinkSheetAdapter:
    def __init__(
        self,
        client: HttpClient,
        *,
        now: Callable[[], datetime] = lambda: datetime.now(UTC),
        workbook_url: str = DEFAULT_URL,
    ) -> None:
        self._client = client
        self._now = now
        self._url = workbook_url
        self._descriptor = validate_source_manifest(
            {
                "manifest_version": 1,
                "provider_id": "world_bank_pink_sheet",
                "adapter_version": "1.0.0",
                "source_kinds": ["economic_series"],
                "authority_class": "official",
                "point_in_time_support": False,
                "allowed_usage_profiles": [
                    "local_personal_research",
                    "internal_research",
                    "commercial",
                ],
                "raw_retention": "permitted",
                "terms_url": "https://www.worldbank.org/en/about/legal/terms-of-use-for-world-bank-websites",
                "terms_reviewed_at": "2026-07-19",
                "independence_group": "world-bank-commodities",
                "geographies": ["GLOBAL"],
                "frequencies": ["monthly"],
                "normal_update_cadence": "monthly",
                "required_attribution": "World Bank Commodity Price Data (The Pink Sheet)",
            }
        )

    def descriptor(self) -> SourceDescriptor:
        return self._descriptor

    def capabilities(self) -> EconomicSeriesCapabilities:
        return EconomicSeriesCapabilities(True, False, frozenset({"monthly"}))

    async def health(self) -> HealthStatus:
        await self._client.get(self._url)
        return HealthStatus(HealthState.HEALTHY, self._now(), "Pink Sheet workbook reachable")

    async def observations(self, request: EconomicSeriesRequest) -> DataEnvelope[EconomicSeries]:
        if request.point_in_time_at is not None:
            raise PointInTimeUnavailableError("Pink Sheet workbook does not expose vintages")
        response = await self._client.get(self._url)
        observations, units = _read_workbook(response.body, request)
        retrieved = self._now()
        return DataEnvelope(
            EconomicSeries(request.series_id, request.geography, request.series_id, observations),
            self._descriptor,
            request.series_id,
            self._url,
            retrieved,
            retrieved,
            units=units,
            content_hash=response.content_hash,
            request_fingerprint=response.request_fingerprint,
            usage_policy_id="world-bank-terms-2026-07-19",
            required_attribution=self._descriptor.required_attribution,
            warnings=("Current historical workbook may contain revisions",),
            missing_fields=("published_at", "vintage", "revised_at"),
        )


def _read_workbook(
    content: bytes, request: EconomicSeriesRequest
) -> tuple[tuple[EconomicObservation, ...], str | None]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise ExternalDataInvalidError("Pink Sheet workbook is invalid") from error
    try:
        sheet = workbook["Monthly Prices"]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    header_index = next(
        (index for index, row in enumerate(rows) if row and str(row[0]).strip() == "Date"),
        None,
    )
    if header_index is None:
        raise ExternalDataInvalidError("Pink Sheet Date header is missing")
    headers = tuple(str(item).strip() if item is not None else "" for item in rows[header_index])
    try:
        column = headers.index(request.series_id)
    except ValueError as error:
        raise ExternalDataInvalidError(
            f"Pink Sheet series is unavailable: {request.series_id}"
        ) from error
    units_row = rows[header_index + 1] if len(rows) > header_index + 1 else ()
    units = str(units_row[column]).strip() if len(units_row) > column else None
    result: list[EconomicObservation] = []
    for row in rows[header_index + 2 :]:
        if len(row) <= column:
            continue
        period = _period(row[0])
        if period is None or not request.start_period <= period <= request.end_period:
            continue
        try:
            value = None if row[column] is None else Decimal(str(row[column]))
        except InvalidOperation as error:
            raise ExternalDataInvalidError("Pink Sheet value is invalid") from error
        result.append(
            EconomicObservation(
                request.series_id,
                request.geography,
                period,
                value,
                units,
            )
        )
    return tuple(result), units


def _period(value: object) -> str | None:
    if isinstance(value, datetime):
        return f"{value.year:04d}-{value.month:02d}"
    if isinstance(value, date):
        return f"{value.year:04d}-{value.month:02d}"
    if isinstance(value, str):
        stripped = value.strip()
        return stripped[:7] if len(stripped) >= 7 else None
    return None
